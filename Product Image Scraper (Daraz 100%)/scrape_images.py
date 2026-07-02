"""
Product Image Scraper
======================
Auto-detects any .xlsx file in the same folder, visits each unique product URL,
extracts up to 5 full-size image links, and writes the updated file.

Usage:
    pip install requests beautifulsoup4 openpyxl pandas lxml selenium
    # For othoba.com / pickaboo.com: also install ChromeDriver matching your Chrome version
    #   https://googlechromelabs.github.io/chrome-for-testing/
    python scrape_images.py
    python scrape_images.py my_catalog.xlsx        # optional: specify file

Input:  any .xlsx file in the same folder (or pass filename as argument)
Output: <original_filename>_updated.xlsx
"""

import sys
import re
import json
import time
import requests
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ─────────────────────────────────────────────────────────────────────
DELAY       = 1.5
MAX_RETRIES = 3
TIMEOUT     = 15
MAX_IMAGES  = 5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

# ── File detection ──────────────────────────────────────────────────────────────

def find_input_file():
    """
    Resolve the input Excel file by priority:
      1. Command-line argument:  python scrape_images.py myfile.xlsx
      2. Only one .xlsx in the script's folder  -> use it
      3. Multiple .xlsx files -> ask user to pick
    """
    script_dir = Path(__file__).parent

    # 1. CLI arg
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.is_absolute():
            p = script_dir / p
        if not p.exists():
            raise FileNotFoundError(f"File not found: {p}")
        return p

    # 2 & 3. Auto-detect (skip *_updated.xlsx files)
    xlsx_files = sorted(
        f for f in script_dir.glob("*.xlsx") if "_updated" not in f.name
    )
    if not xlsx_files:
        raise FileNotFoundError(
            f"No .xlsx files found in {script_dir}.\n"
            "Place your catalog file next to this script, "
            "or pass the filename as an argument:\n"
            "  python scrape_images.py my_catalog.xlsx"
        )
    if len(xlsx_files) == 1:
        print(f"Auto-detected file: {xlsx_files[0].name}")
        return xlsx_files[0]

    # Multiple files – let user choose
    print("Multiple .xlsx files found. Which one should I use?")
    for i, f in enumerate(xlsx_files, 1):
        print(f"  [{i}] {f.name}")
    while True:
        choice = input("Enter number: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(xlsx_files):
            return xlsx_files[int(choice) - 1]
        print("  Invalid choice, try again.")


def find_url_column(df):
    """Return the name of the column that contains product URLs."""
    # Prefer exact match
    for name in ("Product URL", "product_url", "URL", "url", "Link", "link"):
        if name in df.columns:
            return name
    # Fallback: first column whose values start with http
    for col in df.columns:
        sample = df[col].dropna().astype(str)
        if sample.str.startswith("http").mean() > 0.5:
            print(f"  Auto-detected URL column: '{col}'")
            return col
    raise ValueError(
        "Could not find a URL column. "
        "Make sure one column contains product page links starting with https://"
    )

# ── Image helpers ───────────────────────────────────────────────────────────────

SIZE_RE = re.compile(r"-\d+x\d+(?=\.[a-zA-Z]{2,5}$)")

def strip_size(url):
    """Remove WooCommerce thumbnail suffix: image-600x600.jpg -> image.jpg"""
    return SIZE_RE.sub("", url)


def is_product_image(url):
    low = url.lower()
    bad = ("placeholder", "logo", "icon", "favicon", "banner",
           "pixel", "1x1", "spacer", "woocommerce-placeholder")
    return (
        any(low.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp"))
        and not any(b in low for b in bad)
    )


def extract_images_inest(soup, page_url):
    """
    Extractor for inest.com.bd custom storefront.
    Images live at:  https://inest.com.bd//posadmin/images/product/large/<filename>
    The page also exposes the main image in the og:image meta tag.
    Strategy:
      1. Collect every <img src> whose path contains /posadmin/images/product/large/
      2. Fallback: og:image meta tag
      3. Deduplicate; prefer 'large' over 'xsmall'/'small' variants
    """
    found = []
    seen  = set()

    def add(url):
        if not url:
            return
        url = url.strip().replace("//posadmin", "/posadmin")   # fix double-slash
        if url not in seen and is_product_image(url):
            seen.add(url)
            found.append(url)

    # P1 – <img> tags with /posadmin/images/product/large/ in the src
    for img in soup.find_all("img"):
        for attr in ("src", "data-src", "data-lazy-src"):
            src = img.get(attr, "") or ""
            if "/posadmin/images/product/large/" in src:
                # Make sure we have an absolute URL
                if src.startswith("//"):
                    src = "https:" + src
                elif src.startswith("/"):
                    src = "https://inest.com.bd" + src
                add(src)
        if len(found) >= MAX_IMAGES:
            break

    # P2 – og:image meta (usually the main product photo)
    if len(found) < MAX_IMAGES:
        og = soup.find("meta", property="og:image")
        if og and og.get("content"):
            add(og["content"])

    return found[:MAX_IMAGES]


def extract_images_rokomari(soup, page_url):
    """
    Extractor for rokomari.com (book store).
    Only the og:image meta tag is used — it contains the book cover URL
    from rokbucket.rokomari.io, which is the single image needed.
    """
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        url = og["content"].strip()
        if is_product_image(url):
            return [url]
    return []


def extract_images_othoba(soup, page_url):
    """
    Extractor for othoba.com.
    Product images are served from images.othoba.com/images/thumbs/<file>.

    IMPORTANT: The nav menu also loads category banner images from the same
    domain (e.g. 'Global Finds_20.png', 'Quick Commerce_20.png'). These must
    be excluded. The og:image meta tag always points directly to the correct
    product image, so we use only that.
    """
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        url = og["content"].strip()
        if is_product_image(url):
            return [url]
    return []


def extract_images_pickaboo(soup, page_url):
    """
    Extractor for pickaboo.com.
    Product images are served from their CDN (cdn.pickaboo.com or similar).
    og:image is the main product photo; additional gallery images come
    from <img> tags with the CDN domain.
    """
    found = []
    seen  = set()

    def add(url):
        if not url:
            return
        url = url.strip()
        if url.startswith("//"):
            url = "https:" + url
        if url not in seen and is_product_image(url):
            seen.add(url)
            found.append(url)

    # P1 – og:image
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        add(og["content"])

    # P2 – <img> tags pointing to pickaboo CDN
    for img in soup.find_all("img"):
        for attr in ("src", "data-src", "data-lazy-src"):
            src = img.get(attr, "") or ""
            if "pickaboo.com" in src and "/uploads/" in src:
                add(src)
        if len(found) >= MAX_IMAGES:
            break

    return found[:MAX_IMAGES]


# ── Selenium browser fetch (for sites that block requests) ──────────────────────

# Sites that require a real browser to load
BROWSER_REQUIRED_DOMAINS = ("othoba.com", "pickaboo.com")

def fetch_with_browser(url):
    """
    Use Selenium headless Chrome to fetch pages from sites that block requests.
    Requires:  pip install selenium  +  ChromeDriver in PATH
    Returns page HTML as string, or None on failure.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
    except ImportError:
        print("  [!] Selenium not installed. Run:  pip install selenium")
        print("      Also ensure ChromeDriver is in your PATH.")
        return None

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    driver = None
    try:
        driver = webdriver.Chrome(options=options)
        driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        driver.get(url)
        # Wait up to 10s for <body> to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(2)   # let JS render product images
        return driver.page_source
    except Exception as e:
        print(f"  [!] Browser fetch failed: {e}")
        return None
    finally:
        if driver:
            driver.quit()


def extract_images(html, page_url):
    """
    P0 – inest.com.bd custom storefront (detected by domain)
    P0 – rokomari.com book store (detected by domain)
    P0 – othoba.com e-commerce (detected by domain)
    P0 – pickaboo.com e-commerce (detected by domain)
    P1 – WooCommerce variation JSON
    P2 – All <img> tags on the product domain
    P3 – .woocommerce-product-gallery__image img fallback
    """
    soup = BeautifulSoup(html, "lxml")

    # Route to site-specific extractors
    if "inest.com.bd" in page_url:
        return extract_images_inest(soup, page_url)

    if "rokomari.com" in page_url:
        return extract_images_rokomari(soup, page_url)

    if "othoba.com" in page_url:
        return extract_images_othoba(soup, page_url)

    if "pickaboo.com" in page_url:
        return extract_images_pickaboo(soup, page_url)

    found = []
    seen  = set()

    def add(url):
        if not url:
            return
        clean = strip_size(url.strip())
        if clean and clean not in seen and is_product_image(clean):
            seen.add(clean)
            found.append(clean)

    # P1 – variation JSON in <script> blocks
    for script in soup.find_all("script"):
        text = script.string or ""
        if not text:
            continue
        for m in re.finditer(
            r'"(?:full_src|full|src|url)"\s*:\s*"(https?://[^"]+\.[a-zA-Z]{2,5})"', text
        ):
            add(m.group(1))
        if len(found) >= MAX_IMAGES:
            break

    # P1b – data-product_variations attribute
    for tag in soup.find_all(attrs={"data-product_variations": True}):
        try:
            for v in json.loads(tag["data-product_variations"]):
                img = v.get("image", {})
                for k in ("full_src", "src", "url"):
                    add(img.get(k, ""))
        except (json.JSONDecodeError, TypeError):
            pass

    # P2 – all <img> tags on same domain / WP uploads
    if len(found) < MAX_IMAGES:
        domain = re.sub(r"https?://([^/]+).*", r"\1", page_url)
        for img in soup.find_all("img"):
            src = (
                img.get("src") or
                img.get("data-src") or
                img.get("data-lazy-src") or ""
            )
            if domain in src or "/wp-content/uploads/" in src:
                add(src)
            if len(found) >= MAX_IMAGES:
                break

    # P3 – WooCommerce gallery
    if len(found) < MAX_IMAGES:
        for img in soup.select(".woocommerce-product-gallery__image img"):
            add(img.get("src") or img.get("data-src") or "")

    return found[:MAX_IMAGES]


# ── Fetching ────────────────────────────────────────────────────────────────────

def needs_browser(url):
    """Return True if this URL must be fetched with a real browser."""
    return any(domain in url for domain in BROWSER_REQUIRED_DOMAINS)


def fetch_with_retry(session, url):
    # Some sites block plain HTTP requests — use headless Chrome for them
    if needs_browser(url):
        print("  (using browser fetch)")
        return fetch_with_browser(url)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            return r.text
        except requests.RequestException as e:
            print(f"    attempt {attempt}/{MAX_RETRIES} failed: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY * 2)
    return None


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    input_path  = find_input_file()
    output_path = input_path.parent / (input_path.stem + "_updated.xlsx")

    print(f"\nInput  : {input_path.name}")
    print(f"Output : {output_path.name}\n")

    df      = pd.read_excel(input_path)
    url_col = find_url_column(df)

    unique_urls = df[url_col].dropna().unique().tolist()
    cache = {}

    session = requests.Session()
    session.headers.update(HEADERS)

    print(f"Scraping {len(unique_urls)} unique product URLs …\n")

    for i, url in enumerate(unique_urls, 1):
        print(f"[{i}/{len(unique_urls)}] {url}")
        html = fetch_with_retry(session, url)
        if html:
            imgs = extract_images(html, url)
            cache[url] = imgs
            print(f"  -> {len(imgs)} image(s) found")
        else:
            cache[url] = []
            print("  -> failed; leaving blank")
        if i < len(unique_urls):
            time.sleep(DELAY)

    # ── Write Excel ─────────────────────────────────────────────────────────
    print("\nWriting output …")
    wb = load_workbook(input_path)
    ws = wb.active

    img_start_col = ws.max_column + 1

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align   = Alignment(horizontal="left", vertical="center")

    for j in range(1, MAX_IMAGES + 1):
        col  = img_start_col + j - 1
        cell = ws.cell(row=1, column=col, value=f"Image Link {j}")
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 60

    url_col_idx = df.columns.get_loc(url_col) + 1   # 1-based

    for row_idx in range(2, ws.max_row + 1):
        url  = ws.cell(row=row_idx, column=url_col_idx).value
        imgs = cache.get(str(url).strip(), [])
        for j, img_url in enumerate(imgs):
            cell = ws.cell(row=row_idx, column=img_start_col + j, value=img_url)
            cell.alignment = cell_align

    wb.save(output_path)

    # ── Build result data for summary window ────────────────────────────────
    found_urls     = [(u, cache[u]) for u in unique_urls if cache.get(u)]
    not_found_urls = [(u, cache[u]) for u in unique_urls if not cache.get(u)]
    total_rows     = len(df)
    total_unique   = len(unique_urls)
    total_found    = len(found_urls)
    total_missing  = len(not_found_urls)

    print(f"\nDone!  Saved -> {output_path.name}")
    print(f"  Rows total   : {total_rows}")
    print(f"  URLs scraped : {total_unique}")
    print(f"  Found        : {total_found}")
    print(f"  Not found    : {total_missing}")

    show_summary_window(
        output_path    = output_path,
        total_rows     = total_rows,
        total_unique   = total_unique,
        found_urls     = found_urls,
        not_found_urls = not_found_urls,
    )


# ── Summary GUI ──────────────────────────────────────────────────────────────

def show_summary_window(output_path, total_rows, total_unique, found_urls, not_found_urls):
    try:
        import tkinter as tk
        from tkinter import ttk
    except ImportError:
        print("(tkinter not available – skipping popup)")
        return

    total_found   = len(found_urls)
    total_missing = len(not_found_urls)

    root = tk.Tk()
    root.title("Scrape Complete")
    root.resizable(True, True)
    root.configure(bg="#1e1e2e")

    CLR_BG       = "#1e1e2e"
    CLR_PANEL    = "#2a2a3e"
    CLR_GREEN    = "#4caf50"
    CLR_RED      = "#f44336"
    CLR_ACCENT   = "#7c6af7"
    CLR_TEXT     = "#e0e0e0"
    CLR_SUBTEXT  = "#9e9ec8"
    CLR_FOUND_BG = "#1a3a1a"
    CLR_MISS_BG  = "#3a1a1a"
    CLR_ROW_ALT  = "#252535"

    FNT_TITLE = ("Segoe UI", 16, "bold")
    FNT_STAT  = ("Segoe UI", 12, "bold")
    FNT_LABEL = ("Segoe UI", 10)
    FNT_MONO  = ("Consolas", 9)
    FNT_BTN   = ("Segoe UI", 10, "bold")

    # Title bar
    title_frame = tk.Frame(root, bg=CLR_ACCENT, pady=14)
    title_frame.pack(fill="x")
    tk.Label(title_frame, text="Scrape Complete", font=FNT_TITLE,
             bg=CLR_ACCENT, fg="white").pack()
    tk.Label(title_frame, text=str(output_path.name), font=FNT_LABEL,
             bg=CLR_ACCENT, fg="#d0ccff").pack()

    # Stat cards
    stats_frame = tk.Frame(root, bg=CLR_BG, pady=10)
    stats_frame.pack(fill="x", padx=20)
    cards = [
        ("Total Rows",   str(total_rows),    CLR_ACCENT),
        ("URLs Scraped", str(total_unique),   "#2196f3"),
        ("Found",        str(total_found),    CLR_GREEN),
        ("Not Found",    str(total_missing),  CLR_RED),
    ]
    for col_idx, (label, value, colour) in enumerate(cards):
        card = tk.Frame(stats_frame, bg=CLR_PANEL, padx=18, pady=12)
        card.grid(row=0, column=col_idx, padx=8, sticky="nsew")
        stats_frame.columnconfigure(col_idx, weight=1)
        tk.Label(card, text=value, font=("Segoe UI", 22, "bold"),
                 bg=CLR_PANEL, fg=colour).pack()
        tk.Label(card, text=label, font=FNT_LABEL,
                 bg=CLR_PANEL, fg=CLR_SUBTEXT).pack()

    # Notebook tabs
    style = ttk.Style()
    style.theme_use("default")
    style.configure("Dark.TNotebook", background=CLR_BG, borderwidth=0)
    style.configure("Dark.TNotebook.Tab", background=CLR_PANEL, foreground=CLR_SUBTEXT,
                    padding=[14, 6], font=FNT_LABEL)
    style.map("Dark.TNotebook.Tab",
              background=[("selected", CLR_ACCENT)],
              foreground=[("selected", "white")])

    nb = ttk.Notebook(root, style="Dark.TNotebook")
    nb.pack(fill="both", expand=True, padx=20, pady=(6, 0))

    def make_list_tab(parent, items, bg_row):
        frame = tk.Frame(parent, bg=CLR_BG)
        hdr = tk.Frame(frame, bg=CLR_PANEL)
        hdr.pack(fill="x")
        tk.Label(hdr, text="#",    width=5,  anchor="center",
                 font=FNT_STAT, bg=CLR_PANEL, fg=CLR_SUBTEXT).grid(row=0, column=0, padx=4, pady=4)
        tk.Label(hdr, text="Product URL", anchor="w",
                 font=FNT_STAT, bg=CLR_PANEL, fg=CLR_SUBTEXT).grid(row=0, column=1, sticky="w", padx=4)
        tk.Label(hdr, text="Images", width=8, anchor="center",
                 font=FNT_STAT, bg=CLR_PANEL, fg=CLR_SUBTEXT).grid(row=0, column=2, padx=4)
        hdr.columnconfigure(1, weight=1)

        canvas = tk.Canvas(frame, bg=CLR_BG, highlightthickness=0)
        sb     = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        inner  = tk.Frame(canvas, bg=CLR_BG)
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        for idx, (url, imgs) in enumerate(items):
            row_bg    = bg_row if idx % 2 == 0 else CLR_ROW_ALT
            row       = tk.Frame(inner, bg=row_bg)
            row.pack(fill="x")
            row.columnconfigure(1, weight=1)
            img_count = len(imgs)
            icon      = "+" if img_count > 0 else "x"
            iclr      = CLR_GREEN if img_count > 0 else CLR_RED
            tk.Label(row, text=str(idx + 1), width=5, anchor="center",
                     font=FNT_MONO, bg=row_bg, fg=CLR_SUBTEXT).grid(row=0, column=0, padx=4, pady=3)
            tk.Label(row, text=url, anchor="w",
                     font=FNT_MONO, bg=row_bg, fg=CLR_TEXT).grid(row=0, column=1, sticky="w", padx=4)
            tk.Label(row, text=f"{icon} {img_count}", width=8, anchor="center",
                     font=FNT_STAT, bg=row_bg, fg=iclr).grid(row=0, column=2, padx=4)

        def _scroll(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _scroll)
        return frame

    tab_found = make_list_tab(nb, found_urls,     CLR_FOUND_BG)
    tab_miss  = make_list_tab(nb, not_found_urls, CLR_MISS_BG)
    nb.add(tab_found, text=f"  Found ({total_found})  ")
    nb.add(tab_miss,  text=f"  Not Found ({total_missing})  ")

    # Bottom bar
    bottom = tk.Frame(root, bg=CLR_BG, pady=12)
    bottom.pack(fill="x", padx=20)
    tk.Label(bottom, text=f"Saved: {output_path}",
             font=FNT_LABEL, bg=CLR_BG, fg=CLR_SUBTEXT,
             wraplength=500, justify="left").pack(side="left")

    def open_folder():
        import subprocess, platform
        folder = str(output_path.parent)
        if platform.system() == "Windows":
            subprocess.Popen(["explorer", folder])
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", folder])
        else:
            subprocess.Popen(["xdg-open", folder])

    tk.Button(bottom, text="Open Folder", font=FNT_BTN,
              bg=CLR_ACCENT, fg="white", relief="flat",
              padx=14, pady=6, cursor="hand2",
              command=open_folder).pack(side="right", padx=(8, 0))
    tk.Button(bottom, text="Close", font=FNT_BTN,
              bg=CLR_PANEL, fg=CLR_TEXT, relief="flat",
              padx=14, pady=6, cursor="hand2",
              command=root.destroy).pack(side="right")

    root.update_idletasks()
    w, h = 800, 580
    sw   = root.winfo_screenwidth()
    sh   = root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
    root.minsize(600, 420)
    root.mainloop()


if __name__ == "__main__":
    main()